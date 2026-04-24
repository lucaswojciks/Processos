from pathlib import Path

import psycopg2
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from tqdm import tqdm


trecho = "Acesso Aeroporto - Acesso Tapera - Acesso SC-401"
pasta_projeto = "145.26 SINASC SIE Litoral"
revisao = "REV00"

str_where = "WHERE TRECHO_PROJETO IN('Acesso Aeroporto', 'Acesso Tapera', ' Acesso SC-401');"

db_params = {
    "dbname": "p14526",
    "user": "postgres",
    "password": "Projev!@spr16",
    "host": "192.168.1.50",
    "port": "5432",
}

view_name = "projetos.planilha_projeto_tacha_tachao"

BASE_DIR = Path(__file__).resolve().parents[3]
OBJECT_OUTPUT_DIR = BASE_DIR / "src" / "tacha_tachao" / "output"

images_path = Path(r"Z:\Geoprocessamento\16_JPG_PLANILHAS\TODAS_HORIZONTAL\TACHA TACHAO")

path_excel = fr"S:\Projetos\{pasta_projeto}\05 GIS\04 SCRIPTS\03_PLANILHAS\MODELOS\Projeto Sinalizacao Tachas e Tachões.xlsx"

output_excel = OBJECT_OUTPUT_DIR / trecho / revisao / "Planilhas_Projeto" / f"{trecho}_Projeto_Tachas.xlsx"

EMU_PER_PIXEL = 9525


def column_width_to_pixels(width):
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def row_height_to_pixels(height):
    if height is None:
        height = 15
    return int(height * 96 / 72)


def calcular_tamanho_proporcional(img_w, img_h, max_w, max_h, fator_reducao=0.85):
    if not img_w or not img_h:
        return max_w, max_h

    escala = min(max_w / img_w, max_h / img_h)
    escala = min(escala, 1.0)
    escala *= fator_reducao

    nova_w = int(img_w * escala)
    nova_h = int(img_h * escala)

    return max(nova_w, 1), max(nova_h, 1)


def inserir_imagem_centralizada(ws, caminho_imagem, linha, coluna=19, margem_px=4, fator_reducao=0.85):
    if not caminho_imagem or not caminho_imagem.exists():
        ws.cell(row=linha, column=coluna).value = "Imagem não encontrada"
        return

    col_letter = get_column_letter(coluna)

    largura_coluna = ws.column_dimensions[col_letter].width
    altura_linha = ws.row_dimensions[linha].height

    largura_celula_px = column_width_to_pixels(largura_coluna)
    altura_celula_px = row_height_to_pixels(altura_linha)

    max_w = max(largura_celula_px - 2 * margem_px, 1)
    max_h = max(altura_celula_px - 2 * margem_px, 1)

    with PILImage.open(caminho_imagem) as pil_img:
        img_w, img_h = pil_img.size

    nova_w, nova_h = calcular_tamanho_proporcional(
        img_w,
        img_h,
        max_w,
        max_h,
        fator_reducao=fator_reducao,
    )

    img = OpenpyxlImage(str(caminho_imagem))
    img.width = nova_w
    img.height = nova_h

    offset_x_px = max(0, (largura_celula_px - nova_w) // 2)
    offset_y_px = max(0, (altura_celula_px - nova_h) // 2)

    marker = AnchorMarker(
        col=coluna - 1,
        row=linha - 1,
        colOff=offset_x_px * EMU_PER_PIXEL,
        rowOff=offset_y_px * EMU_PER_PIXEL,
    )

    size = XDRPositiveSize2D(
        cx=nova_w * EMU_PER_PIXEL,
        cy=nova_h * EMU_PER_PIXEL,
    )

    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


query = f"""
    SELECT
        id_serial,
        trecho_projeto,
        km_ini,
        lat_ini,
        long_ini,
        km_fim,
        lat_fim,
        long_fim,
        sentido_pista,
        prancha,
        descricao,
        cadencia,
        bordo_dir,
        bordo_esq,
        eixo,
        zpa,
        extensao,
        quantidade,
        cod_imagem,
        id
    FROM {view_name}
    {str_where}
"""


print("1/4 Conectando ao banco e buscando dados...")
conn = psycopg2.connect(**db_params)
cursor = conn.cursor()

try:
    cursor.execute(query)
    dados = cursor.fetchall()

    total_rows = len(dados)
    print(f"   {total_rows} registros encontrados.")

    print("2/4 Carregando modelo Excel...")
    wb = openpyxl.load_workbook(path_excel)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["S"].width = 20
    ws.sheet_view.showGridLines = False

    linha_excel = 8

    print("3/4 Preenchendo planilha...")
    with tqdm(
        total=total_rows,
        desc="Gerando planilha",
        unit="linha",
        ncols=100,
    ) as pbar:
        for item in dados:
            (
                id_serial,
                trecho_projeto,
                km_ini,
                lat_ini,
                long_ini,
                km_fim,
                lat_fim,
                long_fim,
                sentido_pista,
                prancha,
                descricao,
                cadencia,
                bordo_dir,
                bordo_esq,
                eixo,
                zpa,
                extensao,
                quantidade,
                cod_imagem,
                id_objeto,
            ) = item

            caminho_imagem = images_path / cod_imagem if cod_imagem else None

            ws.row_dimensions[linha_excel].height = 80

            ws[f"A{linha_excel}"] = id_serial
            ws[f"B{linha_excel}"] = trecho_projeto
            ws[f"C{linha_excel}"] = km_ini
            ws[f"D{linha_excel}"] = lat_ini
            ws[f"E{linha_excel}"] = long_ini
            ws[f"F{linha_excel}"] = km_fim
            ws[f"G{linha_excel}"] = lat_fim
            ws[f"H{linha_excel}"] = long_fim
            ws[f"I{linha_excel}"] = sentido_pista
            ws[f"J{linha_excel}"] = prancha
            ws[f"K{linha_excel}"] = descricao
            ws[f"L{linha_excel}"] = cadencia
            ws[f"M{linha_excel}"] = bordo_dir
            ws[f"N{linha_excel}"] = bordo_esq
            ws[f"O{linha_excel}"] = eixo
            ws[f"P{linha_excel}"] = zpa
            ws[f"Q{linha_excel}"] = extensao
            ws[f"R{linha_excel}"] = quantidade
            ws[f"T{linha_excel}"] = id_objeto

            try:
                inserir_imagem_centralizada(
                    ws=ws,
                    caminho_imagem=caminho_imagem,
                    linha=linha_excel,
                    coluna=19,  # S
                    margem_px=4,
                    fator_reducao=0.85,
                )
            except Exception as exc:
                tqdm.write(f"Erro ao inserir imagem na linha {linha_excel}: {exc}")

            for col in range(1, 20):
                cell = ws.cell(row=linha_excel, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            linha_excel += 1
            pbar.update(1)

    print("4/4 Salvando arquivo...")
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)
    wb.close()

    print(f"\n✅ Relatório gerado com sucesso em: {output_excel}")

finally:
    cursor.close()
    conn.close()