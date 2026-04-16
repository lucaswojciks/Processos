from pathlib import Path

import psycopg2
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.units import pixels_to_EMU


DB_NAME = "p14526"
TRECHO = "Acesso Aeroporto - Acesso Tapera - Acesso SC-401"
PASTA_PROJETO = "145.26 SINASC SIE Litoral"
REVISAO = "REV00"

BASE_PROJETO = Path(rf"S:\Projetos\{PASTA_PROJETO}")
EXCEL_MODEL_PATH = BASE_PROJETO / "05 GIS" / "04 SCRIPTS" / "03_PLANILHAS" / "MODELOS" / "Projeto Sinalizacao Vertical e Aerea.xlsx"
IMAGES_PATH = BASE_PROJETO / "05 GIS" / "14 PLACAS" / TRECHO / "PNG"
OUTPUT_EXCEL = Path(__file__).resolve().parents[2] / "output" / f"{TRECHO}_projeto_vertical.xlsx"

IMAGE_COLUMN = "Q"
LARGURA_COLUNA_IMAGEM = 18
ALTURA_LINHA_IMAGEM = 80  # points
IMG_MAX_W = 110  # px
IMG_MAX_H = 72   # px

# Aproximações usadas para centralizar a imagem na célula
CELL_WIDTH_PX = 126
CELL_HEIGHT_PX = 106

DB_PARAMS = {
    "dbname": DB_NAME,
    "user": "postgres",
    "password": "Projev!@spr16",
    "host": "192.168.1.50",
    "port": "5432",
}

QUERY = """
    SELECT
        ID,
        ID_SERIAL,
        TRECHO_PROJETO,
        KM,
        SENTIDO_PISTA,
        LADO_VIA,
        LATITUDE,
        LONGITUDE,
        PRANCHA,
        COD,
        COD_DER,
        DIMENSAO,
        AREA,
        TIPO_SUBSTRATO,
        TIPO_PELICULA,
        TIPO_SUPORTE,
        DIMENSAO_SUPORTE,
        SUP_QTD,
        MATERIAL_CONTRATADO_PLACA,
        MATERIAL_CONTRATADO_SUPORTE
    FROM PROJETOS.PLANILHA_PROJETO_VERTICAL;
"""


def redimensionar_imagem(img: OpenpyxlImage, max_w: int, max_h: int) -> OpenpyxlImage:
    """Redimensiona a imagem mantendo a proporção original, sem ampliar imagens pequenas."""
    largura_original = img.width
    altura_original = img.height

    if not largura_original or not altura_original:
        return img

    fator = min(max_w / largura_original, max_h / altura_original, 1.0)

    img.width = int(largura_original * fator)
    img.height = int(altura_original * fator)
    return img


def inserir_imagem(ws, caminho_imagem: Path, celula: str, max_w: int = 110, max_h: int = 72) -> None:
    """Insere a imagem centralizada dentro da célula usando OneCellAnchor."""
    img = OpenpyxlImage(str(caminho_imagem))
    img = redimensionar_imagem(img, max_w, max_h)

    col_letras = "".join(filter(str.isalpha, celula))
    row_num = int("".join(filter(str.isdigit, celula)))

    col_idx = column_index_from_string(col_letras) - 1
    row_idx = row_num - 1

    offset_x_px = max(int((CELL_WIDTH_PX - img.width) / 2), 0)
    offset_y_px = max(int((CELL_HEIGHT_PX - img.height) / 2), 0)

    marker = AnchorMarker(
        col=col_idx,
        colOff=pixels_to_EMU(offset_x_px),
        row=row_idx,
        rowOff=pixels_to_EMU(offset_y_px),
    )

    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(img.width),
        cy=pixels_to_EMU(img.height),
    )

    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def configurar_planilha(ws) -> Border:
    """Aplica configurações iniciais da planilha e retorna a borda padrão."""
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)

    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions[IMAGE_COLUMN].width = LARGURA_COLUNA_IMAGEM
    ws.column_dimensions["T"].hidden = True
    ws.sheet_view.showGridLines = False

    return border


def aplicar_estilo_linha(ws, excel_row: int, border: Border) -> None:
    """Aplica alinhamento e borda em todas as colunas da linha."""
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        cell = ws[f"{col}{excel_row}"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def preencher_linha(ws, excel_row: int, row_data: tuple) -> str:
    """Preenche uma linha da planilha com os dados do banco e retorna o código da imagem."""
    (
        record_id,
        id_serial,
        trecho_projeto,
        km,
        sentido_pista,
        lado_via,
        latitude,
        longitude,
        prancha,
        cod,
        cod_der,
        dimensao,
        area,
        tipo_substrato,
        tipo_pelicula,
        tipo_suporte,
        dimensao_suporte,
        sup_qtd,
        material_contratado_placa,
        material_contratado_suporte,
    ) = row_data

    valores_por_coluna = {
        "A": id_serial,
        "B": trecho_projeto,
        "C": km,
        "D": sentido_pista,
        "E": lado_via,
        "F": latitude,
        "G": longitude,
        "H": prancha,
        "I": cod_der,
        "J": dimensao,
        "K": area,
        "L": tipo_substrato,
        "M": tipo_pelicula,
        "N": tipo_suporte,
        "O": dimensao_suporte,
        "P": sup_qtd,
        "R": material_contratado_placa,
        "S": material_contratado_suporte,
        "T": record_id,
    }

    for coluna, valor in valores_por_coluna.items():
        ws[f"{coluna}{excel_row}"] = valor

    return str(cod)


def main() -> None:
    with psycopg2.connect(**DB_PARAMS) as conn:
        with conn.cursor() as cursor:
            cursor.execute(QUERY)
            vertical_rows = cursor.fetchall()

    total_rows = len(vertical_rows)

    wb = load_workbook(EXCEL_MODEL_PATH)
    ws = wb.active

    border = configurar_planilha(ws)
    excel_row = 8

    for indice, vertical_row in enumerate(vertical_rows, start=1):
        cod = str(vertical_row[9])
        id_serial = vertical_row[1]

        print(f"Processando {indice}/{total_rows} - ID_SERIAL: {id_serial}")

        ws.row_dimensions[excel_row].height = ALTURA_LINHA_IMAGEM

        preencher_linha(ws, excel_row, vertical_row)
        aplicar_estilo_linha(ws, excel_row, border)

        image_path = IMAGES_PATH / f"{cod}.png"
        if image_path.exists():
            try:
                inserir_imagem(
                    ws=ws,
                    caminho_imagem=image_path,
                    celula=f"Q{excel_row}",
                    max_w=IMG_MAX_W,
                    max_h=IMG_MAX_H,
                )
            except Exception as exc:
                print(f"Erro ao inserir imagem {image_path}: {exc}")

        excel_row += 1

    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_EXCEL)

    print(f"\n✅ Relatório gerado com sucesso em: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()