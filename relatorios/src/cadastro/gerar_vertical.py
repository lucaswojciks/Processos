from pathlib import Path
from shutil import copyfile

import psycopg2
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, TwoCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.units import pixels_to_EMU
from tqdm import tqdm


DB_NAME = "p14526"
TRECHO = "Acesso Aeroporto - Acesso Tapera - Acesso SC-401"
PASTA_PROJETO = "145.26 SINASC SIE Litoral"
REVISAO = "REV00"

BASE_PROJETO = Path(rf"S:\Projetos\{PASTA_PROJETO}")
IMAGES_PATH = Path(rf"\\192.168.1.50\Projetos\Projeview\{DB_NAME}")
ICONS_PATH = Path(r"\\192.168.1.50\Projetos\Projeview\_sistema\object_data\placa\icons")
EXCEL_MODEL_PATH = BASE_PROJETO / "05 GIS" / "04 SCRIPTS" / "03_PLANILHAS" / "MODELOS" / "Cadastro Sinalizacao Vertical.xlsx"
OUTPUT_EXCEL = Path(__file__).resolve().parents[2] / "output" / f"{TRECHO}_cadastro_vertical.xlsx"
VERTICAL_PATH = Path(__file__).resolve().parents[2] / "output" / "VERTICAL"

IMAGE_ITEM_COLUMN = "L"
IMAGE_ICON_COLUMN = "M"
STATUS_COLUMN = "N"
ID_COLUMN = "O"

IMAGE_ITEM_COLUMN_WIDTH = 18
IMAGE_ICON_COLUMN_WIDTH = 12
IMAGE_ROW_HEIGHT = 80  # points

IMG_WIDTH = 121
IMG_HEIGHT = 99
ICON_MAX_W = 60
ICON_MAX_H = 60
EMU_PER_PIXEL = 9525

DB_PARAMS = {
    "dbname": DB_NAME,
    "user": "postgres",
    "password": "Projev!@spr16",
    "host": "192.168.1.50",
    "port": "5432",
}

QUERY = """
    SELECT
        ID,
        ID_SERIAL,
        TRECHO_PROJETO,
        KM,
        LADO_VIA,
        LATITUDE,
        LONGITUDE,
        PRANCHA,
        COD,
        COD_DER,
        SUPORTE,
        DIMENSAO,
        AREA,
        IMAGEM,
        SITUACAO
    FROM PROJETOS.PLANILHA_CADASTRO_VERTICAL;
"""


def redimensionar_imagem(img: OpenpyxlImage, max_w: int, max_h: int) -> OpenpyxlImage:
    """Redimensiona a imagem mantendo a proporção original, sem ampliar imagens pequenas."""
    largura_original = img.width
    altura_original = img.height

    if not largura_original or not altura_original:
        return img

    fator = min(max_w / largura_original, max_h / altura_original, 1.0)

    img.width = int(largura_original * fator)
    img.height = int(altura_original * fator)
    return img


def configurar_planilha(ws) -> Border:
    """Aplica configurações iniciais da planilha e retorna a borda padrão."""
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.column_dimensions["A"].hidden = False
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions[IMAGE_ITEM_COLUMN].width = IMAGE_ITEM_COLUMN_WIDTH
    ws.column_dimensions[IMAGE_ICON_COLUMN].width = IMAGE_ICON_COLUMN_WIDTH
    ws.column_dimensions[ID_COLUMN].hidden = True
    ws.sheet_view.showGridLines = False
    ws["C1"] = "Cadastro - Sinalização Vertical"

    return border


def aplicar_estilo_linha(ws, excel_row: int, border: Border) -> None:
    """Aplica alinhamento e borda em todas as colunas da linha."""
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        cell = ws[f"{col}{excel_row}"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def obter_nome_imagem(imagem: dict | None, situacao: str) -> str | None:
    """Retorna o nome da imagem correspondente à situação."""
    if not imagem or not isinstance(imagem, dict):
        return None

    return imagem.get(situacao)


def inserir_imagem_objeto(ws, column: str, caminho_imagem: Path, excel_row: int) -> None:
    """Insere a imagem do objeto com TwoCellAnchor e tamanho fixo."""
    img = OpenpyxlImage(str(caminho_imagem))
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT

    col_idx_zero = column_index_from_string(column) - 1
    row_idx_zero = excel_row - 1

    largura_coluna = ws.column_dimensions[column].width
    if largura_coluna is None:
        largura_coluna = 8.43

    altura_linha = ws.row_dimensions[excel_row].height
    if altura_linha is None:
        altura_linha = 15

    largura_px = int(largura_coluna * 7)
    altura_px = int(altura_linha * 96 / 72)

    sobra_x = max(0, largura_px - img.width)
    sobra_y = max(0, altura_px - img.height)

    offset_x_px = sobra_x // 2
    offset_y_px = sobra_y // 2

    from_marker = AnchorMarker(
        col=col_idx_zero,
        row=row_idx_zero,
        colOff=offset_x_px * EMU_PER_PIXEL,
        rowOff=offset_y_px * EMU_PER_PIXEL,
    )

    to_marker = AnchorMarker(
        col=col_idx_zero,
        row=row_idx_zero,
        colOff=(offset_x_px + img.width) * EMU_PER_PIXEL,
        rowOff=(offset_y_px + img.height) * EMU_PER_PIXEL,
    )

    img.anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=from_marker,
        to=to_marker,
    )

    ws.add_image(img)


def inserir_icone(ws, column: str, caminho_imagem: Path, excel_row: int, max_w: int = 60, max_h: int = 60) -> None:
    """Insere o ícone centralizado na célula usando OneCellAnchor."""
    img = OpenpyxlImage(str(caminho_imagem))
    img = redimensionar_imagem(img, max_w, max_h)

    col_idx = column_index_from_string(column) - 1
    row_idx = excel_row - 1

    largura_coluna = ws.column_dimensions[column].width
    if largura_coluna is None:
        largura_coluna = 8.43

    altura_linha = ws.row_dimensions[excel_row].height
    if altura_linha is None:
        altura_linha = 15

    cell_w_px = int(largura_coluna * 7)
    cell_h_px = int(altura_linha * 96 / 72)

    offset_x_px = max(int((cell_w_px - img.width) / 2), 0)
    offset_y_px = max(int((cell_h_px - img.height) / 2), 0)

    marker = AnchorMarker(
        col=col_idx,
        colOff=pixels_to_EMU(offset_x_px),
        row=row_idx,
        rowOff=pixels_to_EMU(offset_y_px),
    )

    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(img.width),
        cy=pixels_to_EMU(img.height),
    )

    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def copiar_imagem_para_saida(caminho_imagem: Path) -> None:
    """Copia a imagem para a pasta de saída, se ainda não existir."""
    VERTICAL_PATH.mkdir(parents=True, exist_ok=True)

    destino_img = VERTICAL_PATH / caminho_imagem.name
    if not destino_img.exists():
        copyfile(caminho_imagem, destino_img)


def preencher_linha(ws, excel_row: int, row_data: tuple) -> None:
    """Preenche uma linha da planilha."""
    (
        record_id,
        id_serial,
        trecho_projeto,
        km,
        lado_via,
        latitude,
        longitude,
        prancha,
        cod,
        cod_der,
        suporte,
        dimensao,
        area,
        imagem,
        situacao,
    ) = row_data

    valores_por_coluna = {
        "A": id_serial,
        "B": trecho_projeto,
        "C": km,
        "D": lado_via,
        "E": latitude,
        "F": longitude,
        "G": prancha,
        "H": cod_der,
        "I": suporte,
        "J": dimensao,
        "K": area,
        STATUS_COLUMN: situacao,
        ID_COLUMN: record_id,
    }

    for coluna, valor in valores_por_coluna.items():
        ws[f"{coluna}{excel_row}"] = valor


def main() -> None:
    print("1/4 Conectando ao banco e buscando dados...")
    with psycopg2.connect(**DB_PARAMS) as conn:
        with conn.cursor() as cursor:
            cursor.execute(QUERY)
            placas_rows = cursor.fetchall()

    total_rows = len(placas_rows)
    print(f"   {total_rows} registros encontrados.")

    print("2/4 Carregando modelo Excel...")
    wb = load_workbook(EXCEL_MODEL_PATH)
    ws = wb.active

    border = configurar_planilha(ws)
    excel_row = 8

    print("3/4 Preenchendo planilha...")
    with tqdm(
        total=total_rows,
        desc="Gerando planilha",
        unit="linha",
        ncols=100,
    ) as pbar:
        for row_data in placas_rows:
            imagem = row_data[13]
            situacao = row_data[14]
            cod = row_data[8]

            if imagem is None:
                pbar.update(1)
                continue

            imagem_name = obter_nome_imagem(imagem, situacao)
            if imagem_name is None:
                pbar.update(1)
                continue

            ws.row_dimensions[excel_row].height = IMAGE_ROW_HEIGHT

            preencher_linha(ws, excel_row, row_data)
            aplicar_estilo_linha(ws, excel_row, border)

            image_path = IMAGES_PATH / imagem_name
            if image_path.exists():
                try:
                    inserir_imagem_objeto(ws, IMAGE_ITEM_COLUMN, image_path, excel_row)
                    copiar_imagem_para_saida(image_path)
                except Exception as exc:
                    tqdm.write(f"Erro ao adicionar ou copiar imagem {image_path.name}: {exc}")
            else:
                tqdm.write(f"Imagem não encontrada: {image_path.name}")

            icon_path = ICONS_PATH / f"{cod}.png"
            if icon_path.exists():
                try:
                    inserir_icone(
                        ws=ws,
                        column=IMAGE_ICON_COLUMN,
                        caminho_imagem=icon_path,
                        excel_row=excel_row,
                        max_w=ICON_MAX_W,
                        max_h=ICON_MAX_H,
                    )
                except Exception as exc:
                    tqdm.write(f"Erro ao adicionar ícone {icon_path.name}: {exc}")
            else:
                tqdm.write(f"Ícone não encontrado: {icon_path.name}")

            excel_row += 1
            pbar.update(1)

    print("4/4 Salvando arquivo...")
    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_EXCEL)
    wb.close()

    print(f"\n✅ Relatório gerado com sucesso em: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()